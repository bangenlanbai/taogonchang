# -*- coding:utf-8  -*-
# @Time     : 2022/8/4 22:38
# @Author   : BGLB
# @Software : PyCharm
import json
import os
import time
import traceback
from contextlib import closing
from functools import reduce

import requests
from openpyxl import load_workbook

from core import TaoGongChang

header = {
    'authority': 'tgc.tmall.com',
    'method': 'POST',
    'scheme': 'https',
    'accept': 'application/json, text/plain, */*',
    'accept-language': 'zh-CN,zh;q=0.9',
    'bx-v': '2.2.2',
    'cache-control': 'no-cache',
    'content-type': 'application/json;charset=UTF-8',
    'cookie': 'isg=BKKiHm4ENBBBmijMp_Lb5qT58ygE86YNAOF3U-w7xZXAv0I51IExHFo97_tDrx6l; tfstk=cD55B-x300mSKrrX0La4LB5NrVRGZwg6R8tlPo4dcsGYgn75i_lwfmkGKDdeJE1..; l=eB_xNpImL7RudKr2BOfZourza77TjIRAguPzaNbMiOCPO7fw5MmlW6xI2rLeCnGVhs1kR3oSoXzWBeYBqIfYLsbr42FJXhMmn; X-XSRF-TOKEN=b0b5879c-43d2-429c-8fc0-8da151558de8; SCMBIZTYPE=176000; SCMSESSID=1d73bed2af1f78ed227ecbdaab1f8a99@HAVANA; cookie17=UUpgR1XIK6lQS2vrBQ%3D%3D; _nk_=scm09620215; SCMLOCALE=zh-cn; uc3=nk2=EF2TYziLgGExNCs%3D&lg2=V32FPkk%2Fw0dUvg%3D%3D&vt3=F8dCv4GxOwITPc80CIE%3D&id2=UUpgR1XIK6lQS2vrBQ%3D%3D; xlly_s=1; _tb_token_=96b613d1d1f1; uc4=id4=0%40U2gqyOiLaxMX6umQUmlDNZLOALUOFRet&nk4=0%40EoTGj1Z%2B2%2FffunOpCnbotEXVT7L0ew%3D%3D; cookie2=1d73bed2af1f78ed227ecbdaab1f8a99; csg=237b5ddf; t=da2eb272c88bef428a82acbd13b6ad8b; lid=scm09620215; XSRF-TOKEN=8b1658c2-803d-46da-9e7f-a896dbca2ffa; tracknick=scm09620215; sgcookie=E100BrnqLOZ%2BDv8QWYjLUUVwndrnemFjbku5yybcQ4M9GnUP%2FpFgohSQCGgXm4ltPOfaVmBK%2BLEg%2BuM4lBijce%2FfHbFcmZLhsnXg0IIpbpx1by0%3D; lgc=scm09620215; cna=WKl0Gy/QpgQCAbfvnGKQllUw',
    'origin': 'https://tgc.tmall.com',
    'pragma': 'no-cache',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 ' \
                  'Safari/537.36',
    'x-xsrf-token': '8b1658c2-803d-46da-9e7f-a896dbca2ffa',
}


def get_excel():
    """

    :return:
    """
    url = 'https://tgc.tmall.com/api/v1/orderNew/tradeOrderDownload.htm?spm=a26364.20167850.0.0.59664840JWIIAA&_input_charset=utf-8&taskId=40332923&c2mNavigatorShellPage=2&c2mNavigatorPageOpener=1'
    a = request_download_big_file({'url': url, 'headers': header}, './data.csv')
    print(a)


def request_download_big_file(request_kwargs, local_path, ):
    try:
        if os.path.exists(local_path):
            os.remove(local_path)
        local_path_tmp = local_path+'.tmp'
        param = {'url': '', 'stream': True, 'headers': ''}
        param.update(request_kwargs)
        with closing(requests.get(**param)) as r:
            # r = requests.get(url=file_url, verify=False, stream=True)
            if r.status_code == 200:
                with open(local_path_tmp, "wb") as f:
                    # f.write(r.content)
                    for chunk in r.iter_content(chunk_size=4096):
                        if chunk:
                            f.write(chunk)
            else:
                return False
        os.rename(local_path_tmp, local_path)
    except Exception:
        print(traceback.format_exc())
        return False
    return True


def get_qOsi(mainOrder):
    url = 'https://tgc.tmall.com/ds/api/v1/o/qOsi'
    param = {"mainOrderId": '2802160872802626426', "infoKeys": ["buyerNick", "fullName", "mobilephone", "fullAddress"]}
    rep = requests.post(url, json=param, headers=header)
    result = rep.json()

    if result.get('success'):
        return True, result.get('data')
    return False, result.get('errorMessage')


def get_order():
    url = 'https://tgc.tmall.com/api/v1/orderNew/getTradeOrders.htm'
    order_list = []
    page = 1
    pageSize = 10
    param = {
        'pageNo': page,
        'pageSize': pageSize,
        'sourceTradeId': '',
        'status': 'PAID'
    }
    rep = requests.get(url, params=param, headers=header)
    total = rep.json().get('paginator', {}).get('total')
    if total > pageSize:
        while True:
            rep = requests.get(url, params=param, headers=header)
            if rep.json().get('success'):
                order = rep.json().get('data')
                for item in order:
                    print(item)
                    order_list.extend(item.get('detailOrders'))

            if len(order_list) >= total:
                break
            param['pageNo'] = page+1
    else:
        order = rep.json().get('data')
        for item in order:
            order_list.append(item.get('detailOrders'))
    run_function = lambda x, y: x if y in x else x+[y]
    order_list = reduce(run_function, [[], ]+order_list)

    return order_list


def read_excel(path):
    """

    :return:
    """
    res = []
    try:
        wb = load_workbook(path)
        sheet = wb.active
        n = 1
        for row in sheet.iter_rows(values_only=True, min_row=2):
            n += 1
            mainOrder = row[1]
            if mainOrder:
                flag, data = get_qOsi(mainOrder)
                print(f'{mainOrder}: {data}')
                if flag and data:
                    sheet.cell(row=n, column=3, value=data.get('buyerNick'))
                    sheet.cell(row=n, column=12, value=data.get('fullName'))
                    sheet.cell(row=n, column=13, value=data.get('fullAddress'))
                    sheet.cell(row=n, column=14, value=data.get('mobilephone'))
                # time.sleep(.1)
        wb.save('订单数据.xlsx')
    except Exception:
        print(traceback.format_exc())
        return False, traceback.format_exc()
    return True, res

def save_excel(data):
    """
        保存数据
    :param data:
    :return:
    """
    try:
        tmpl_path = os.path.join('tmpl.xlsx')
        wb = load_workbook(tmpl_path)
        sheet = wb.active
        for index, order in enumerate(data):
            order_id = order.get('sourceTradeId')
            index = index+2
            flag, data = get_qOsi(order_id)
            print(f'{order_id}: {data}')
            order.update(data)
            # sheet.cell(row=index, column=1, value=order_id)  # 快递公司
            # sheet.cell(row=index, column=2, value=order_id)  # 快递单号
            sheet.cell(row=index, column=3, value=order_id)  # 订单编号
            # sheet.cell(row=index, column=4, value=order_id)  # 订单来源
            sheet.cell(row=index, column=5, value=order.get('buyerNick'))  # 买家昵称
            sheet.cell(row=index, column=6, value=order.get('fullName'))  # 收货人姓名
            sheet.cell(row=index, column=7, value=order.get('mobilephone'))  # 收货人手机号
            sheet.cell(row=index, column=8, value=order.get('prov'))  # 省
            sheet.cell(row=index, column=9, value=order.get('city'))  # 市
            sheet.cell(row=index, column=10, value=order.get('area'))  # 区/县
            sheet.cell(row=index, column=11, value=order.get('town'))  # 街道地址
            sheet.cell(row=index, column=12, value=order.get('fullAddress'))  # 详细信息
            # sheet.cell(row=index, column=13, value=order.get('mobilephone'))  # 卖家备注
            # sheet.cell(row=index, column=14, value=order.get('mobilephone'))  # 买家留言
            # sheet.cell(row=index, column=15, value=order.get('mobilephone'))  # 实付金额
            sheet.cell(row=index, column=16, value=order.get('auctionTitle'))  # 商品标题
            sheet.cell(row=index, column=17, value=order.get('outerIdSku'))  # 商家编码
            sheet.cell(row=index, column=18, value=order.get('buyAmount'))  # 商品数量
        wb.save('订单数据_{}_{}.xlsx'.format('aa', time.strftime('%Y-%m-%d_%H_%M_%S')))
    except Exception:
        print("生成订单数据失败\n{}".format(traceback.format_exc()))
        return False

def save_data(data):
    """

    :return:
    """
    with open('./data.json', encoding='utf8', mode='r') as f:
        data = json.load(f)
    save_excel(data)


def test():
    all_task = [TaoGongChang, ]
    for task in all_task:
        task.test()


if __name__ == '__main__':
    # a = get_order()
    # save_data(a)
    # print(len(a))
    # print(get_qOsi(''))
    get_order()
