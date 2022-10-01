# -*- coding:utf-8  -*-
# @Time     : 2022/9/18 17:14
# @Author   : BGLB
# @Software : PyCharm
import os
import stat
import sys
import time
import traceback
import zipfile
from contextlib import closing
from copy import copy
from functools import reduce

import requests
from loguru import logger
from openpyxl import load_workbook
from openpyxl.cell import Cell
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

import json

pre_start_js = """
"""

BASE_DIR = os.path.realpath(os.path.dirname(sys.argv[0]))

PLATFORM_CHOICE = (('tmall', '淘工厂'),)


class Base(object):
    """
        Base
    """

    def __init__(self, login_id, password, platform=None):
        """ init """
        self.login_id = login_id
        self.password = password
        self.platform = platform
        self.platform_str = dict(PLATFORM_CHOICE).get(self.platform)
        self.cookie = ''
        self.driver_path = 'chromedriver.exe'
        self.driver: webdriver.Chrome
        self.browser_user_data = os.path.join(BASE_DIR, 'user_data', self.login_id)
        self.__init_log()
        self.__init_dir()

    def __init_log(self):
        for root, dirs, files in os.walk(os.path.join(BASE_DIR, 'log'), topdown=False):
            for name in files:
                os.remove(os.path.join(root, name))
            for name in dirs:
                os.rmdir(os.path.join(root, name))

        logger.add(os.path.join(BASE_DIR, 'log', f'{self.login_id}.log'), **{'backtrace': True,
                                                                             'diagnose': True,
                                                                             'enqueue': True,
                                                                             'catch': True},
                   level='INFO')
        logger.add(os.path.join(BASE_DIR, 'log', f'error.log'), **{'backtrace': True,
                                                                   'diagnose': True,
                                                                   'enqueue': True,
                                                                   'catch': True},
                   level='ERROR')

        self.log = logger

    def __init_dir(self):
        """

        :return:
        """
        self.tmpl_path = os.path.join(BASE_DIR, f'{self.platform}_tmpl.xlsx')
        self.excel_save_path = os.path.join(
            '{}_订单数据_{}_{}.xlsx'.format(self.platform_str, self.login_id, time.strftime('%Y-%m-%d_%H%M%S')))
        self.data_save_dir = os.path.join(BASE_DIR, 'json', self.platform, self.login_id)
        self.cookie_dir = os.path.join(BASE_DIR, 'cookies', self.platform)
        self.cookie_path = os.path.join(self.cookie_dir, f'{self.login_id}.txt')
        self.data_save_path = os.path.join(self.data_save_dir, f'{self.platform}_'
                                                               f'{time.strftime("%Y-%m-%d_%H_%M_%S")}.json')
        os.makedirs(self.cookie_dir, exist_ok=True)

    def save_cookie(self, cookie_str):
        """

        :return:
        """
        with open(self.cookie_path, mode='w', encoding='utf8') as f:
            f.write(cookie_str)

    def load_cookie(self):
        """

        :return:
        """
        try:
            with open(self.cookie_path, mode='r', encoding='utf8') as f:
                cookie = f.read()
                return str(cookie)
        except Exception:
            return False

    def close_some_server(self):
        """
            关闭浏览器
        :return:
        """
        if self.driver:
            try:
                self.driver.quit()
                self.driver.close()
            except Exception:
                self.log.error('close driver 失败')

    def get_version_via_com(self, file_name):
        import win32api
        try:
            info = win32api.GetFileVersionInfo(file_name, os.sep)
            ms = info['FileVersionMS']
            ls = info['FileVersionLS']
            version = '%d.%d.%d.%d'%(win32api.HIWORD(ms), win32api.LOWORD(ms), win32api.HIWORD(ls), win32api.LOWORD(ls))
            return version
        except:
            return ''

    def get_chrome_version(self, is_windows=True):
        if is_windows:

            path = os.environ['LOCALAPPDATA']+r'\Google\Chrome\Application\chrome.exe'
            path1 = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
            path2 = r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe'
            plist = [path, path1, path2]
            for p in plist:
                v = self.get_version_via_com(p)
                if v:
                    return '.'.join(v.split('.')[:-1])
            return ''
        else:
            version = os.popen("google-chrome-stable --version").read()
            version_num = ".".join(str(version.split(' ')[2]).split('.')[:-1])
            return version_num

    def get_driver_version(self, driver_path: str):
        """
            获取driver_version
        :return:
        """
        if not os.path.exists(driver_path):
            try:
                os.makedirs(os.sep.join(driver_path.split(os.sep)[:-1]), exist_ok=True)
            except Exception:
                pass
            return "0.0.0"
        try:
            if not driver_path.endswith('exe'):
                """linux 系统更改文件权限"""
                os.chmod(driver_path, stat.S_IXGRP)

            outstd = os.popen('{} --version'.format(driver_path)).read()
            # self.log.info(outstd)
            version = outstd.split(' ')[1]
            version = ".".join(version.split(".")[:-1])
            return version
        except Exception as e:
            return "0.0.0"

    def download_driver(self, chrome_version: str, driver_dir: str, is_windows: bool):
        """
            下载与chrome 匹配的driver
        :param chrome_version:
        :param driver_dir:
        :param is_windows:
        :return: bool
        """
        base_url = 'http://npm.taobao.org/mirrors/chromedriver/'
        url = "{}LATEST_RELEASE_{}".format(base_url, chrome_version)
        last_version = requests.get(url).text
        # 下载chromedriver
        if is_windows:
            download_url = "{}{}/chromedriver_win32.zip".format(base_url, last_version)
        else:
            download_url = "{}{}/chromedriver_linux64.zip".format(base_url, last_version)
        file = requests.get(download_url)
        chromedriver_zip_path = os.path.join(driver_dir, 'chromedriver.zip')

        # 保存zip
        with open(chromedriver_zip_path, 'wb') as zip_file:
            zip_file.write(file.content)

        # 解压
        with zipfile.ZipFile(chromedriver_zip_path, 'r') as f:
            for file in f.namelist():
                f.extract(file, driver_dir)
        # self.log.info(chromedriver_zip_path)
        os.remove(chromedriver_zip_path)
        if not is_windows:
            chromeDriver = os.path.join(driver_dir, 'chromedriver')
            os.chmod(chromeDriver, stat.S_IXGRP)

    def check_update_driver(self, driver_path) -> (bool, str):
        """
            检查升级driver
        :return:
        """
        is_windows = driver_path.endswith('exe')
        chrome_version = self.get_chrome_version(is_windows)

        driver_version = self.get_driver_version(driver_path)
        # self.log.info('chrome_version:{}, driver_version:{}'.format(chrome_version, driver_version))
        if chrome_version == driver_version:
            # self.log.info('无需升级')
            return True, ''
        try:
            driver_dir = os.sep.join(driver_path.split(os.sep)[:-1])

            # self.log.info('开始升级chromedriver: 【{} -> {}】'.format(driver_version, chrome_version))
            self.download_driver(chrome_version, driver_dir, is_windows)
            driver_version = self.get_driver_version(driver_path)
            # self.log.info('chromedriver下载安装成功,当前版本【{}】'.format(driver_version))
            if driver_version == chrome_version:
                return True, ''
            else:
                return True, ''
        except Exception as e:
            self.log.info('chromedriver:【{}】下载失败: {}'.format(chrome_version, e))
            return True, traceback.format_exc()

    def driver_init(self, proxy: dict = None, is_phone=False):
        """
        初始化chrome driver
        :type  proxy: dict
        :param proxy: 代理配置 dict{'ip': '_ip', 'port': '_port'}
        :return: webdriver.Chrome()

        """
        options = webdriver.ChromeOptions()
        pres = {'credentials_enable_service': False, 'profile.password_manager_enabled': False}
        options.add_argument('--disable-gpu')
        options.add_argument("--disable-blink-features=AutomationControlled")  # 88版本过检测
        options.add_argument('lang=zh_CN.UTF-8')  # 设置语言
        options.add_argument('--disable-infobars')  # 除去“正受到自动测试软件的控制”
        # options.add_argument("--auto-open-devtools-for-tabs") # 相当于 F12
        # options.add_extension('')  # 添加插件
        if is_phone:
            options.add_experimental_option('mobileEmulation', {'deviceName': 'iPhone X'})  # 模拟iPhone X浏览
        options.add_experimental_option('excludeSwitches', ['enable-automation', 'enable-logging'])  # 过检测
        options.add_experimental_option('useAutomationExtension', False)
        options.add_experimental_option('prefs', pres)  # 禁用保存密码弹框
        options.add_argument(f"user-data-dir={self.browser_user_data}")
        # 添加代理
        if proxy:
            options.add_argument("--proxy-server=http://{}:{}".format(proxy['ip'], proxy['port']))

        if not self.driver_path.endswith('exe'):
            # options.add_argument('--headless')  # 浏览器不提供可视化页面. linux下如果系统不支持可视化不加这条会启动失败
            options.add_argument('--no-sandbox')
        s = Service(executable_path=self.driver_path)
        self.driver = webdriver.Chrome(service=s, options=options)

        self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": pre_start_js
        })
        self.driver.maximize_window()

    def save_data(self, content):
        """

        :param content:
        :return:
        """
        pass

    def read_data(self, path, encoding='utf8'):
        """

        :param path:
        :return:
        """
        with open(path, mode='r', encoding=encoding) as f:
            content = f.read()
        try:
            json_data = json.loads(content)
            return json_data
        except Exception:
            self.log.warning(f'文件[{path}] json格式化失败')
            return content

    @staticmethod
    def request_download_big_file(request_kwargs, local_path, ):
        try:
            if os.path.exists(local_path):
                os.remove(local_path)
            local_path_tmp = local_path+'.tmp'
            param = {'url': '', 'stream': True, 'headers': ''}
            param.update(request_kwargs)
            with closing(requests.get(**param)) as r:
                # r = requests.get(url=file_url, verify=False, stream=True)
                if r.status_code == 200:
                    with open(local_path_tmp, "wb") as f:
                        # f.write(r.content)
                        for chunk in r.iter_content(chunk_size=4096):
                            if chunk:
                                f.write(chunk)
                else:
                    return False
            os.rename(local_path_tmp, local_path)
        except Exception:
            print(traceback.format_exc())
            return False
        return True

    @staticmethod
    def set_excel_cell_style(cell: Cell, temp_cell: Cell):
        """
            获取 某个的所有样式
        :param cell:
        :param temp_cell:
        :return:
        """
        style = copy(temp_cell._style)
        border = copy(temp_cell.border)
        cell._style = style
        cell.border = border

    @staticmethod
    def start():
        """

        :return:
        """
        pass


class TaoGongChang(Base):
    """
        TaoGongChang
    """

    def __init__(self, login_id, password):
        """

        """
        super().__init__(login_id, password, 'tmall')
        self.url_dict = {
            'order': 'https://tgc.tmall.com/api/v1/orderNew/getTradeOrders.htm',
            'qOsi': 'https://tgc.tmall.com/ds/api/v1/o/qOsi',
            'login': 'https://tgc.tmall.com/ds/page/supplier/order-manage'
        }
        self.xsrf_token = '846c7565-c106-46bb-a6e0-51411ad3ba5e'
        self.header = {
            'authority': 'tgc.tmall.com',
            'method': 'POST',
            'scheme': 'https',
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'zh-CN,zh;q=0.9',
            'bx-v': '2.2.2',
            'content-type': 'application/json;charset=UTF-8',
            'cookie': '',
            'origin': 'https://tgc.tmall.com',
            'pragma': 'no-cache',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
            'x-xsrf-token': self.xsrf_token,
        }

    def get_order(self):
        url = 'https://tgc.tmall.com/api/v1/orderNew/getTradeOrders.htm'
        order_list = []
        page = 1
        pageSize = 10
        param = {
            'pageNo': page,
            'pageSize': pageSize,
            'sourceTradeId': '',
            'status': 'PAID'
        }
        rep = requests.get(url, params=param, headers=self.header)
        # print(rep)
        total = rep.json().get('paginator', {}).get('total')
        self.log.info('{}-未发货订单总数：{}'.format(self.login_id, total))
        if total > pageSize:
            while True:
                rep = requests.get(url, params=param, headers=self.header)
                if rep.json().get('success'):
                    order = rep.json().get('data')
                    for item in order:
                        order_list.extend(item.get('detailOrders'))

                if len(order_list) >= total:
                    break
                param['pageNo'] += 1
        else:
            order = rep.json().get('data')
            for item in order:
                order_list.extend(item.get('detailOrders'))
        run_function = lambda x, y: x if y in x else x+[y]
        order_list = reduce(run_function, [[], ]+order_list)
        with open(self.data_save_path, encoding='utf8', mode='w') as f:
            json.dump(order_list, f, ensure_ascii=False)
        return order_list

    def get_qOsi(self, mainOrderId):
        """

        :return:
        """
        url = 'https://tgc.tmall.com/ds/api/v1/o/qOsi'
        param = {"mainOrderId": mainOrderId, "infoKeys": ["buyerNick", "fullName", "mobilephone", "fullAddress"]}
        rep = requests.post(url, json=param, headers=self.header)
        result = rep.json()
        if result.get('success'):
            return True, result.get('data')
        return False, {}

    def login(self):
        """

        :return:
        """
        is_login = self.check_login()
        if is_login:
            return True
        update_result = self.check_update_driver(self.driver_path)
        if not update_result:
            self.log.info('chromedriver升级失败， 任务结束')
            return False
        try:
            self.driver_init()
            login_url = self.url_dict.get('login')
            self.driver.get(login_url)
            self.driver.refresh()
            time.sleep(2)
            if 'login' not in self.driver.current_url:
                self.log.info('账号{}已登录'.format(self.login_id))
                # return True
            else:
                # ele = (By.XPATH, '//*[@id="root"]/div/div[1]/div/div/div/div/a')
                # WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(ele))
                # self.driver.find_element(*ele).click()
                iframe = (By.XPATH, '//*[@id="alibaba-login-box"]')
                WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(iframe))
                self.driver.switch_to.frame(self.driver.find_element(*iframe))

                ele = (By.XPATH, '//*[@id="fm-login-id"]')
                WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(ele))
                self.driver.find_element(*ele).send_keys(self.login_id)
                time.sleep(2)
                ele = (By.XPATH, '//*[@id="fm-login-password"]')
                WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(ele))
                self.driver.find_element(*ele).send_keys(self.password)
                # self.driver.find_element(By.XPATH, '//*[@id="login-form"]/div[6]/button').click()
                while True:
                    if 'login' not in self.driver.current_url:
                        break
                    self.log.info('请手动登录！')
                    time.sleep(5)
            self.driver.refresh()
            self.driver.get(self.url_dict.get('login'))
            time.sleep(2)
            cookie = {}
            for c in self.driver.get_cookies():
                cookie.update({c['name']: c['value']})
                if c['name'] == 'XSRF-TOKEN':
                    self.xsrf_token = c['value']
            self.cookie_dict = cookie
            self.cookie = '; '.join([f'{k}={v}' for k, v in cookie.items()])
            self.save_cookie(self.cookie)
            self.header.update({'cookie': self.cookie})
            return True
        except Exception:
            self.log.error(traceback.format_exc())
            return False
        finally:
            self.close_some_server()

    def check_login(self):
        cookie = self.load_cookie()
        if not cookie:
            return False
        self.header.update({'cookie': cookie, 'x-xsrf-token': ''})

        try:
            rep = requests.get('https://scm.tmall.com/loginStatus', headers=self.header).json()
            print(rep)
            if rep.get('success') and rep.get('data'):
                self.cookie_dict = {}
                self.cookie = cookie
                for item in cookie.split('; '):
                    k, v = item.split('=', maxsplit=1)
                    self.cookie_dict[k] = v
                    if k == 'XSRF-TOKEN':
                        self.xsrf_token = v
                    self.header.update({'x-xsrf-token': v})
                self.log.info('平台：{} 账号: {} 已经登录成功'.format(self.platform_str, self.login_id))
                return True
            else:
                return False
        except Exception:
            self.log.info(traceback.format_exc())
            return False

    def get_excel(self):
        """
            获取excel
        :return:
        """
        try:
            url = 'https://tgc.tmall.com/ds/page/supplier/order-manage'
            self.driver.get(url)
            ele = (By.XPATH, '//div[@class="component-table-batch-operations"]//button')
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(ele))
            self.driver.find_element(*ele).click()
            time.sleep(2)
            self.driver.get('https://tgc.tmall.com/ds/page/supplier/order-download')

            ele = '//*[@id="root"]/div/div[2]/div[1]/table/tbody/tr[1]/td[9]/div/p/a'
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(ele))
            url = self.driver.find_element(*ele).get_attribute('href')
            url += '&c2mNavigatorShellPage=2&c2mNavigatorPageOpener=1'
            self.request_download_big_file({'url': url, 'headers': self.header}, 'data.xlsx')
        except Exception:
            self.log.error('下载表格失败， {}'.format(traceback.format_exc()))
            return False

    def save_excel(self, data, qQsi=False):
        """
            保存数据
        :param data:
        :return:
        """
        try:

            wb = load_workbook(self.tmpl_path)
            sheet = wb.active
            style_index = 2
            for index, order in enumerate(data):
                order_id = order.get('sourceTradeId')
                index = index+2
                sheet.row_dimensions[index].height = 20  # 设置行高

                if qQsi:
                    flag, data = self.get_qOsi(order_id)
                    self.log.info(f'{order_id}: {data}')
                    order.update(data)
                # sheet.cell(row=index, column=1, value=order_id)  # 快递公司
                # sheet.cell(row=index, column=2, value=order_id)  # 快递单号
                sheet_item = sheet.cell(row=index, column=3, value=order_id)  # 订单编号
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=3))

                # sheet.cell(row=index, column=4, value=order_id)  # 订单来源
                sheet_item = sheet.cell(row=index, column=5, value=order.get('buyerNick'))  # 买家昵称
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=5))

                sheet_item = sheet.cell(row=index, column=6, value=order.get('fullName'))  # 收货人姓名
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=6))

                sheet_item = sheet.cell(row=index, column=7, value=order.get('mobilephone'))  # 收货人手机号
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=7))

                sheet_item = sheet.cell(row=index, column=8, value=order.get('prov'))  # 省
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=8))

                sheet_item = sheet.cell(row=index, column=9, value=order.get('city'))  # 市
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=9))

                sheet_item = sheet.cell(row=index, column=10, value=order.get('area'))  # 区/县
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=10))

                sheet_item = sheet.cell(row=index, column=11, value=order.get('town'))  # 街道地址
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=11))

                sheet_item = sheet.cell(row=index, column=12, value=order.get('address'))  # 详细信息
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=12))

                # sheet.cell(row=index, column=13, value=order.get('mobilephone'))  # 卖家备注
                # sheet.cell(row=index, column=14, value=order.get('mobilephone'))  # 买家留言
                # sheet.cell(row=index, column=15, value=order.get('mobilephone'))  # 实付金额
                sheet_item = sheet.cell(row=index, column=16, value=order.get('auctionTitle'))  # 商品标题
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=16))

                sheet_item = sheet.cell(row=index, column=17, value=order.get('outerIdSku'))  # 商家编码
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=17))

                sheet_item = sheet.cell(row=index, column=18, value=order.get('buyAmount'))  # 商品数量
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=18))

                sheet_item = sheet.cell(row=index, column=19, value=order.get('auctionId'))  # 商品ID
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=19))

                shop_info = order.get('orderSkuAttrVOs')
                shop_info_str = ''
                for item in shop_info:
                    item_info = '{}: {}\n'.format(item.get('attrType', "").strip(), item.get('attrValue', "").strip())
                    shop_info_str += item_info

                sheet_item = sheet.cell(row=index, column=20, value=shop_info_str)  # 商品描述
                self.set_excel_cell_style(sheet_item, sheet.cell(row=style_index, column=20))

            wb.save(self.excel_save_path)
        except Exception:
            self.log.error("生成订单数据失败\n{}".format(traceback.format_exc()))
            return False
        return True

    @staticmethod
    def start():
        login_id = input('请输入账号: ')
        password = input('请输入密码: ')
        taoGongChang = TaoGongChang(login_id.strip(), password.strip())
        print('平台: {} 账号: [{}] 开始登录'.format(taoGongChang.platform_str, taoGongChang.login_id))
        try:
            result = taoGongChang.login()
            if not result:
                print('登录失败')
                return
            print('登录成功！开始获取订单信息')
            result = taoGongChang.get_order()

            if result:
                if taoGongChang.save_excel(result, True):
                    print('执行完毕')
                else:
                    print('保存文件失败')
            else:
                print('账号{}: 没有需要发货的订单！'.format(login_id))
                return
        except Exception:
            taoGongChang.log.error(traceback.format_exc())
        finally:
            # taoGongChang.close_some_server()
            os.system('taskkill /f /im chromedriver.exe')

    @staticmethod
    def test():
        # login_id = input('请输入账号: ')
        # password = input('请输入密码: ')
        login_id = '18660998382'
        password = '123456'
        taoGongChang = TaoGongChang(login_id.strip(), password.strip())
        taoGongChang.data_save_path = 'json/tmall/18660998382/data.json'
        print('平台: {} 账号: [{}] 开始登录'.format(taoGongChang.platform_str, taoGongChang.login_id))
        try:
            result = taoGongChang.read_data(taoGongChang.data_save_path)
            if result:
                if taoGongChang.save_excel(result):
                    print('执行完毕')
                else:
                    print('保存文件失败')
            else:
                print('账号{}: 没有需要发货的订单！'.format(login_id))
                return
        except Exception:
            taoGongChang.log.error(traceback.format_exc())
        finally:
            # taoGongChang.close_some_server()
            os.system('taskkill /f /im chromedriver.exe')


class PinDuoDuo(Base):
    """
        PinDuoDuo
    """

    def __init__(self, login_id, password):
        """

        """
        super().__init__(login_id, password, 'pinduoduo')
        self.driver: webdriver.Chrome

    @staticmethod
    def start():
        login_id = input('请输入账号: ')
        password = input('请输入密码: ')
        pinDuoDuo = PinDuoDuo(login_id.strip(), password.strip())
        print('账号: [{}] 开始登录'.format(pinDuoDuo.platform_str, pinDuoDuo.login_id))
        try:
            result = pinDuoDuo.login()
            if not result:
                print('登录失败')
                return
            print('登录成功！开始获取订单信息')
            result = pinDuoDuo.get_order()

            if result:
                if pinDuoDuo.save_excel(result):
                    print('执行完毕')
                else:
                    print('保存文件失败')
            else:
                print('账号{}: 没有需要发货的订单！'.format(login_id))
                return
        except Exception:
            pinDuoDuo.log.error(traceback.format_exc())
        finally:
            # taoGongChang.close_some_server()
            os.system('taskkill /f /im chromedriver.exe')


def main():
    os.system('taskkill /f /im chromedriver.exe')
    print('***********************操作说明***************************'.replace('*', ' '))
    print('********************第一步: 输入账号***********************'.replace('*', ' '))
    print('********************第二步: 输入密码***********************'.replace('*', ' '))
    print('********************第三步: 手动登录***********************'.replace('*', ' '))
    print('********************第四步: 等待程序自己退出*****************'.replace('*', ' '))
    print('*********************************************************'.replace('*', ' '))
    print('***********************注意事项****************************'.replace('*', ' '))
    print('********************1. 不要手动关闭浏览器*******************'.replace('*', ' '))
    print('********************2. 不要手闭本窗口**********************'.replace('*', ' '))
    print('********************3. 必须安装谷歌浏览器*******************'.replace('*', ' '))
    print('********************下面开始执行*******************'.replace('*', ' '))
    time.sleep(2)
    # import ctypes
    # kernel32 = ctypes.windll.kernel32
    # kernel32.SetConsoleMode(kernel32.GetStdHandle(-10), 128)
    all_task = [TaoGongChang, ]
    for task in all_task:
        task.start()
