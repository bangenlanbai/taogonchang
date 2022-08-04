# -*- coding:utf-8  -*-
# @Time     : 2022/8/4 22:38
# @Author   : BGLB
# @Software : PyCharm
import os
import stat
import time
import traceback
import zipfile
from contextlib import closing

import requests
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait


class TaoGongChang(object):
    """
        TaoGongChang
    """

    def __init__(self, login_id, password):
        """

        """
        self.cookie = ''
        self.url_dict = {
            'order': 'https://tgc.tmall.com/api/v1/orderNew/getTradeOrders.htm',
            'get_excel': '',
            'qOsi': 'https://tgc.tmall.com/ds/api/v1/o/qOsi',
            'login': 'https://tgc.tmall.com/'
        }
        self.xsrf_token = '846c7565-c106-46bb-a6e0-51411ad3ba5e'
        self.header = {
            'authority': 'tgc.tmall.com',
            'method': 'POST',
            'scheme': 'https',
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'zh-CN,zh;q=0.9',
            'bx-v': '2.2.2',
            'cache-control': 'no-cache',
            'content-type': 'application/json;charset=UTF-8',
            'cookie': 'cna=xXW2GqIv9RMCAbfvnGIq63vN; xlly_s=1; sgcookie=E100Qj86ID%2Fk0TrRWldMLBRIThghQTZGrW5na7sraTtAUpWD60g6oElStEvTvbC4iQIYjlctaQVdq9mffqVWdfGCud2BzCvMQYNnpd6vBQE2V%2Fs%3D; t=070b3f95a641ba0ca93d5bf37b525788; csg=94bb2a3c; _tb_token_=fef7e593eb6e3; cookie2=1aa13d513e4c871dfb77f7ab5001eef9; SCMLOCALE=zh-cn; _nk_=scm09620215; cookie17=UUpgR1XIK6lQS2vrBQ%3D%3D; SCMSESSID=1aa13d513e4c871dfb77f7ab5001eef9@HAVANA; SCMBIZTYPE=176000; X-XSRF-TOKEN=c39a6889-0197-406b-85c1-23acd3e78714; XSRF-TOKEN=846c7565-c106-46bb-a6e0-51411ad3ba5e; l=eB_izLpnL70Cr9-ABOfwhurza77OMIRfguPzaNbMiOCPOefWREFNW6xiTKLXCnGVnst6R3Wrj_IwBPTEGyznh3v4Gd3hJvSzqdTh.; tfstk=cQDGBNifhfPs-nnr0Aw6nVbtp5kcZC64rxkILhiqGCNUk-kFir5FaRdpxPYW7s1..; isg=BKmpjBp67zOSR9OHa2XjKRzauFUDdp2o7NeytUueWBDZEsgkk8creXBE1LYkijXg',
            'origin': 'https://tgc.tmall.com',
            'pragma': 'no-cache',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 ' \
                          'Safari/537.36',
            'x-xsrf-token': self.xsrf_token,
        }
        self.driver: webdriver.Chrome
        self.driver_path = 'chromedriver.exe'
        self.browser_pre_start_js = './pre_start.js'
        self.login_id = login_id
        self.password = password

    def get_order(self, params: dict):
        """
            获取订单信息
        :param params:
        :return:
        """
        url = self.url_dict.get('order')
        param = {
            'pageNo': 1,
            'pageSize': 10,
            'sourceTradeId': '',
        }
        param = param.update(params)
        rep = requests.get(url, params=param, headers=self.header)
        return rep.json()

    def get_qOsi(self, mainOrderId):
        """

        :return:
        """
        url = 'https://tgc.tmall.com/ds/api/v1/o/qOsi'
        param = {"mainOrderId": mainOrderId, "infoKeys": ["buyerNick", "fullName", "mobilephone", "fullAddress"]}
        rep = requests.post(url, json=param, headers=self.header)
        result = rep.json()
        if result.get('success'):
            return False, {}
        return True, result.get('data')

    def login(self):
        """

        :return:
        """

        update_result = self.check_update_driver(self.driver_path)
        if not update_result:
            print('chromedriver升级失败， 任务结束')
            return
        try:
            self.driver_init()
            login_url = self.url_dict.get('login')
            self.driver.get(login_url)
            self.driver.refresh()
            if 'login' not in self.driver.current_url:
                print('账号{}已登录'.format(self.login_id))
                return True
            else:

                ele = (By.XPATH, '//*[@id="root"]/div/div[1]/div/div/div/div/a')
                WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(ele))
                self.driver.find_element(*ele).click()
                iframe = (By.XPATH, '//*[@id="alibaba-login-box"]')
                WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(iframe))
                self.driver.switch_to.frame(self.driver.find_element(*iframe))

                ele = (By.XPATH, '//*[@id="fm-login-id"]')
                WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(ele))
                self.driver.find_element(*ele).send_keys(self.login_id)
                time.sleep(2)
                ele = (By.XPATH, '//*[@id="fm-login-password"]')
                WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(ele))
                self.driver.find_element(*ele).send_keys(self.password)
                # self.driver.find_element(By.XPATH, '//*[@id="login-form"]/div[6]/button').click()
                while True:
                    if 'login' not in self.driver.current_url:
                        break
                    print('请手动登录！')
                    time.sleep(20)
            cookie = {}
            for c in self.driver.get_cookies():
                cookie.update({c['name']: c['value']})
                if c['name'] == 'XSRF-TOKEN':
                    self.xsrf_token = c['value']
            self.cookie = cookie
            self.header.update({'cookie': cookie})
            return True
        except Exception:
            print(traceback.format_exc())
            return False
        finally:
            self.close_some_server()

    def get_version_via_com(self, file_name):
        import win32api
        try:
            info = win32api.GetFileVersionInfo(file_name, os.sep)
            ms = info['FileVersionMS']
            ls = info['FileVersionLS']
            version = '%d.%d.%d.%d'%(win32api.HIWORD(ms), win32api.LOWORD(ms), win32api.HIWORD(ls), win32api.LOWORD(ls))
            return version
        except:
            return ''

    def get_chrome_version(self, is_windows=True):
        if is_windows:

            path = os.environ['LOCALAPPDATA']+r'\Google\Chrome\Application\chrome.exe'
            path1 = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
            path2 = r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe'
            plist = [path, path1, path2]
            for p in plist:
                v = self.get_version_via_com(p)
                if v:
                    return '.'.join(v.split('.')[:-1])
            return ''
        else:
            version = os.popen("google-chrome-stable --version").read()
            version_num = ".".join(str(version.split(' ')[2]).split('.')[:-1])
            return version_num

    def get_driver_version(self, driver_path: str):
        """
            获取driver_version
        :return:
        """
        if not os.path.exists(driver_path):
            try:
                os.makedirs(os.sep.join(driver_path.split(os.sep)[:-1]), exist_ok=True)
            except Exception:
                pass
            return "0.0.0"
        try:
            if not driver_path.endswith('exe'):
                """linux 系统更改文件权限"""
                os.chmod(driver_path, stat.S_IXGRP)

            outstd = os.popen('{} --version'.format(driver_path)).read()
            # print(outstd)
            version = outstd.split(' ')[1]
            version = ".".join(version.split(".")[:-1])
            return version
        except Exception as e:
            return "0.0.0"

    def download_driver(self, chrome_version: str, driver_dir: str, is_windows: bool):
        """
            下载与chrome 匹配的driver
        :param chrome_version:
        :param driver_dir:
        :param is_windows:
        :return: bool
        """
        base_url = 'http://npm.taobao.org/mirrors/chromedriver/'
        url = "{}LATEST_RELEASE_{}".format(base_url, chrome_version)
        last_version = requests.get(url).text
        # 下载chromedriver
        if is_windows:
            download_url = "{}{}/chromedriver_win32.zip".format(base_url, last_version)
        else:
            download_url = "{}{}/chromedriver_linux64.zip".format(base_url, last_version)
        file = requests.get(download_url)
        chromedriver_zip_path = os.path.join(driver_dir, 'chromedriver.zip')

        # 保存zip
        with open(chromedriver_zip_path, 'wb') as zip_file:
            zip_file.write(file.content)

        # 解压
        with zipfile.ZipFile(chromedriver_zip_path, 'r') as f:
            for file in f.namelist():
                f.extract(file, driver_dir)
        # print(chromedriver_zip_path)
        os.remove(chromedriver_zip_path)
        if not is_windows:
            chromeDriver = os.path.join(driver_dir, 'chromedriver')
            os.chmod(chromeDriver, stat.S_IXGRP)

    def check_update_driver(self, driver_path) -> (bool, str):
        """
            检查升级driver
        :return:
        """
        is_windows = driver_path.endswith('exe')
        chrome_version = self.get_chrome_version(is_windows)

        driver_version = self.get_driver_version(driver_path)
        # print('chrome_version:{}, driver_version:{}'.format(chrome_version, driver_version))
        if chrome_version == driver_version:
            # print('无需升级')
            return True, ''
        try:
            driver_dir = os.sep.join(driver_path.split(os.sep)[:-1])

            # print('开始升级chromedriver: 【{} -> {}】'.format(driver_version, chrome_version))
            self.download_driver(chrome_version, driver_dir, is_windows)
            driver_version = self.get_driver_version(driver_path)
            # print('chromedriver下载安装成功,当前版本【{}】'.format(driver_version))
            if driver_version == chrome_version:
                return True, ''
            else:
                return True, ''
        except Exception as e:
            print('chromedriver:【{}】下载失败: {}'.format(chrome_version, e))
            return True, traceback.format_exc()

    def driver_init(self, proxy: dict = None, is_phone=False):
        """
        初始化chrome driver
        :type  proxy: dict
        :param proxy: 代理配置 dict{'ip': '_ip', 'port': '_port'}
        :return: webdriver.Chrome()

        """
        options = webdriver.ChromeOptions()
        pres = {'credentials_enable_service': False, 'profile.password_manager_enabled': False}
        options.add_argument('--disable-gpu')
        options.add_argument("--disable-blink-features=AutomationControlled")  # 88版本过检测
        options.add_argument('lang=zh_CN.UTF-8')  # 设置语言
        options.add_argument('--disable-infobars')  # 除去“正受到自动测试软件的控制”
        # options.add_argument("--auto-open-devtools-for-tabs") # 相当于 F12
        # options.add_extension('')  # 添加插件
        if is_phone:
            options.add_experimental_option('mobileEmulation', {'deviceName': 'iPhone X'})  # 模拟iPhone X浏览
        options.add_experimental_option('excludeSwitches', ['enable-automation', 'enable-logging'])  # 过检测
        options.add_experimental_option('useAutomationExtension', False)
        options.add_experimental_option('prefs', pres)  # 禁用保存密码弹框
        options.add_argument("user-data-dir=./user_data/")
        # 添加代理
        if proxy:
            options.add_argument("--proxy-server=http://{}:{}".format(proxy['ip'], proxy['port']))

        if not self.driver_path.endswith('exe'):
            # options.add_argument('--headless')  # 浏览器不提供可视化页面. linux下如果系统不支持可视化不加这条会启动失败
            options.add_argument('--no-sandbox')
        s = Service(executable_path=self.driver_path)
        self.driver = webdriver.Chrome(service=s, options=options)

        # 屏蔽浏览器中的window.navigator.webdriver = true
        with open(self.browser_pre_start_js) as f:
            source_js = f.read()
        # self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
        #                             {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"})
        self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": source_js
        })
        self.driver.maximize_window()

    def get_excel(self):
        """
            获取excel
        :return:
        """
        try:
            url = 'https://tgc.tmall.com/ds/page/supplier/order-manage'
            self.driver.get(url)
            ele = (By.XPATH, '//div[@class="component-table-batch-operations"]//button')
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(ele))
            self.driver.find_element(*ele).click()
            time.sleep(2)
            self.driver.get('https://tgc.tmall.com/ds/page/supplier/order-download')

            ele = '//*[@id="root"]/div/div[2]/div[1]/table/tbody/tr[1]/td[9]/div/p/a'
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(ele))
            url = self.driver.find_element(*ele).get_attribute('href')
            url += '&c2mNavigatorShellPage=2&c2mNavigatorPageOpener=1'
            self.request_download_big_file({'url': url, 'headers': self.header}, 'data.xlsx')
        except Exception:
            print('下载表格失败， {}'.format(traceback.format_exc()))
            return False

    def update_excel(self):
        """

        :return:
        """
        try:
            wb = load_workbook('./data.xlsx')
            sheet = wb.active
            n = 1
            for row in sheet.iter_rows(values_only=True, min_row=2):
                n += 1
                mainOrder = row[1]
                if mainOrder:
                    flag, data = self.get_qOsi(mainOrder)
                    print(f'{mainOrder}: {data}')
                    if flag and data:
                        sheet.cell(row=n, column=3, value=data.get('buyerNick'))
                        sheet.cell(row=n, column=12, value=data.get('fullName'))
                        sheet.cell(row=n, column=13, value=data.get('fullAddress'))
                        sheet.cell(row=n, column=14, value=data.get('mobilephone'))
                    # time.sleep(.1)
            wb.save('订单数据.xlsx')
        except Exception:
            print("填充表格失败\n{}".format(traceback.format_exc()))
            return False
        return True

    @staticmethod
    def request_download_big_file(request_kwargs, local_path, ):
        try:
            if os.path.exists(local_path):
                os.remove(local_path)
            local_path_tmp = local_path+'.tmp'
            param = {'url': '', 'stream': True, 'headers': ''}
            param.update(request_kwargs)
            with closing(requests.get(**param)) as r:
                # r = requests.get(url=file_url, verify=False, stream=True)
                if r.status_code == 200:
                    with open(local_path_tmp, "wb") as f:
                        # f.write(r.content)
                        for chunk in r.iter_content(chunk_size=4096):
                            if chunk:
                                f.write(chunk)
                else:
                    return False
            os.rename(local_path_tmp, local_path)
        except Exception:
            print(traceback.format_exc())
            return False
        return True

    def close_some_server(self):
        """
            关闭浏览器
        :return:
        """
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                print('close driver 失败')


def main():
    taoGongChang = TaoGongChang('18298892447', 'abc123abcd1234')
    result = taoGongChang.login()
    if not result:
        print('登录失败')
    result = taoGongChang.get_excel()

    if result:
        if taoGongChang.update_excel():
            print('执行完毕')
            return True


if __name__ == '__main__':
    main()
